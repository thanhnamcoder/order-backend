import os
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

folder_path = r"Data\POG_DATA"
output_file = r"Data\POG_DATA\merged_result.xlsx"

def copy_sheet_contents(src_ws, tgt_ws):
    """Copy values, basic styles, column widths, row heights and merged cells
    from src_ws into tgt_ws. This avoids openpyxl's copy_worksheet restriction
    which only works within the same Workbook.
    """
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        try:
            if dim.width is not None:
                tgt_ws.column_dimensions[col_letter].width = dim.width
        except Exception:
            # ignore any weird column keys
            pass

    # Copy row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        try:
            if getattr(dim, 'height', None) is not None:
                tgt_ws.row_dimensions[row_idx].height = dim.height
        except Exception:
            pass

    # Copy merged cells
    if src_ws.merged_cells.ranges:
        for merged in src_ws.merged_cells.ranges:
            try:
                tgt_ws.merge_cells(str(merged))
            except Exception:
                pass

    # Copy cell values and basic styles
    from openpyxl.cell.cell import MergedCell

    for row in src_ws.iter_rows():
        for cell in row:
            # skip MergedCell when trying to access attributes not available
            if isinstance(cell, MergedCell):
                # still set value into the top-left of merged range is handled by non-merged cell
                continue
            # get column index safely
            col_idx = getattr(cell, 'col_idx', None)
            if col_idx is None:
                # fallback: try to parse from column letter
                try:
                    col_idx = column_index_from_string(cell.column)
                except Exception:
                    col_idx = None

            # if we couldn't determine column index, skip this cell
            if col_idx is None:
                continue

            new_cell = tgt_ws.cell(row=cell.row, column=col_idx, value=cell.value)
            # Copy a few common style attributes (if present)
            try:
                if cell.has_style:
                    if cell.font is not None:
                        new_cell.font = copy.copy(cell.font)
                    if cell.border is not None:
                        new_cell.border = copy.copy(cell.border)
                    if cell.fill is not None:
                        new_cell.fill = copy.copy(cell.fill)
                    if cell.number_format is not None:
                        new_cell.number_format = copy.copy(cell.number_format)
                    if cell.protection is not None:
                        new_cell.protection = copy.copy(cell.protection)
                    if cell.alignment is not None:
                        new_cell.alignment = copy.copy(cell.alignment)
            except Exception:
                # If any style copy fails, fall back to value-only copy
                pass


merged_wb = Workbook()
# remove default sheet if present
try:
    merged_wb.remove(merged_wb.active)
except Exception:
    pass

for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        try:
            wb = load_workbook(file_path, data_only=True)
            src_ws = wb.active  # lấy sheet đầu tiên

            # create a new sheet in merged_wb and copy contents
            sheet_title = os.path.splitext(filename)[0][:31]
            tgt_ws = merged_wb.create_sheet(title=sheet_title)
            copy_sheet_contents(src_ws, tgt_ws)
            print(f"Đã sao chép sheet từ: {filename} -> {sheet_title}")
        except Exception as e:
            print(f"Lỗi khi xử lý {filename}: {e}")

# === Tạo index để tìm kiếm toàn bộ workbook ===
index_rows = []  # list of (sheet, cell, value)
for ws in merged_wb.worksheets:
    # skip any future generated helper sheets
    if ws.title in ("Find", "FindIndex"):
        continue
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            try:
                val = cell.value
            except Exception:
                val = None
            if val is None:
                continue
            # convert to string for search
            s = str(val)
            index_rows.append((ws.title, cell.coordinate, s))

if index_rows:
    # create helper index sheet
    if 'FindIndex' in merged_wb.sheetnames:
        idx_ws = merged_wb['FindIndex']
    else:
        idx_ws = merged_wb.create_sheet('FindIndex')

    # write header
    idx_ws.append(['Sheet', 'Cell', 'Value'])
    for r in index_rows:
        idx_ws.append(list(r))

    # create visible Find sheet with input and formulas
    if 'Find' in merged_wb.sheetnames:
        find_ws = merged_wb['Find']
    else:
        find_ws = merged_wb.create_sheet('Find')

    # input label and cell
    find_ws['A1'] = 'Search'
    find_ws['B1'] = ''  # user types search term here
    find_ws['A2'] = 'First match (Sheet)'
    # INDEX/MATCH (first occurrence) - works on older Excel
    find_ws['B2'] = "=IF($B$1=\"\",\"\",IFERROR(INDEX(FindIndex!A:A,MATCH(\"*\"&$B$1&\"*\",FindIndex!C:C,0)),\"No match\"))"
    find_ws['A3'] = 'First match (Cell)'
    find_ws['B3'] = "=IF($B$1=\"\",\"\",IFERROR(INDEX(FindIndex!B:B,MATCH(\"*\"&$B$1&\"*\",FindIndex!C:C,0)),\"\"))"
    find_ws['A4'] = 'First match (Value)'
    find_ws['B4'] = "=IF($B$1=\"\",\"\",IFERROR(INDEX(FindIndex!C:C,MATCH(\"*\"&$B$1&\"*\",FindIndex!C:C,0)),\"\"))"
    # Add a clickable hyperlink to jump to the first match
    # C2 will be a Go link that jumps to the sheet and cell returned in B2/B3
    find_ws['C1'] = 'Action'
    # create a hyperlink to open the PDF with the same base name in the PDF folder (relative to Data/POG_DATA)
    # e.g. PDF/<sheetname>.pdf
    find_ws['C2'] = '=IF($B$2="","",HYPERLINK("PDF/" & $B$2 & ".pdf", "Open PDF"))'

    # Dynamic full-results using FILTER (Office 365). This will show all matches as a spill array.
    # Place header for full results
    find_ws['A6'] = 'Sheet'
    find_ws['B6'] = 'Cell'
    find_ws['C6'] = 'Value'
    # Put FILTER formula starting at A7
    # The FILTER uses SEARCH to find substring matches (case-insensitive behavior depends on locale)
    find_ws['A7'] = "=IF($B$1=\"\",\"\",IFERROR(FILTER(FindIndex!A:C,ISNUMBER(SEARCH($B$1,FindIndex!C:C))),\"No matches\"))"
    # best-effort helper column for 'Go' links next to the FILTER results (works for first row; users can copy down if needed)
    find_ws['D6'] = 'Go'
    # helper 'Open' link for the first filtered result row that opens the PDF file
    find_ws['D7'] = '=IF(A7="","",HYPERLINK("PDF/" & A7 & ".pdf", "Open"))'

    # hide the index sheet to keep workbook tidy
    try:
        idx_ws.sheet_state = 'hidden'
    except Exception:
        pass

    # set Find as active sheet so workbook opens there
    try:
        merged_wb.active = merged_wb.sheetnames.index('Find')
    except Exception:
        pass

merged_wb.save(output_file)
print(f"✅ Đã gộp xong (giá trị + định dạng cơ bản). File: {output_file}")
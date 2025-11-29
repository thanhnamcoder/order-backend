import requests
import json
from HandHealdLogin import *
import os
import openpyxl
from login_utils import get_login_info
def run_get_data_pog(file_type="excel"):
    """
    file_type: 'excel' (default) hoặc 'pdf'.
    Nếu là excel thì tải file .xlsx, nếu pdf thì tải file .pdf
    """
    result = {
        "status": "success",
        "downloaded_files": [],
        "item_code_count": 0,
        "item_code_json": None,
        "errors": []
    }
    try:
        store_code = get_login_info("store_cd")
        # 1) LOGIN
        session = login(
            url_login="http://storeportal.circlek.com.vn:82/HHT/Login",
            username=store_code,
            password="123123"
        )

        # 2) GET Menu
        get_menu(session)

        # 3) GET token từ Planogram
        token = get_planogram_token(session)

        # 4) POST Planogram View (bắt buộc để server load dữ liệu)
        post_planogram_view(session, token)

        # 5) POST Planogram_Read → TRẢ VỀ DATA CHUẨN
        data = post_planogram_read(session)

        # 6) Lọc danh sách file theo loại
        ext = ".xlsx" if file_type == "excel" else ".pdf"
        file_ids = [
            item["ID"]
            for item in data.get("Data", [])
            if item.get("FileName") and item["FileName"].lower().endswith(ext)
        ]

        # 7) Tạo thư mục lưu file
        save_dir = os.path.join("Data", "POG_DATA", "EXCEL" if file_type == "excel" else "PDF")
        json_dir = os.path.join("Data", "POG_DATA", "JSON")
        os.makedirs(save_dir, exist_ok=True)

        downloaded_files = []

        # 8) Tải từng file
        for item in data.get("Data", []):
            if item.get("FileName") and item["FileName"].lower().endswith(ext):
                file_id = item["ID"]
                file_name = item["FileName"]
                file_url = f"http://storeportal.circlek.com.vn:82/HHT/Mers/ViewFile?ID={file_id}"

                file_resp = session.get(file_url)
                if file_resp.status_code == 200:
                    file_path = os.path.join(save_dir, file_name)
                    with open(file_path, "wb") as f:
                        f.write(file_resp.content)
                    downloaded_files.append(file_path)
                else:
                    result["errors"].append(f"Tải file thất bại cho ID {file_id}, status {file_resp.status_code}")

        result["downloaded_files"] = downloaded_files

        # Nếu là excel thì trích xuất item_code
        if file_type == "excel":
            def extract_item_code_to_json(file_paths, output_json):
                all_values = set()
                for file in file_paths:
                    try:
                        wb = openpyxl.load_workbook(file, data_only=True)
                        ws = wb.active
                        item_code_col = None
                        # Tìm cột Item Code
                        for row in ws.iter_rows(min_row=1, max_row=10):
                            for cell in row:
                                if (
                                    cell.value
                                    and str(cell.value).strip().lower() in ["item code", "itemcode", "item_code"]
                                ):
                                    item_code_col = cell.column
                                    break
                            if item_code_col:
                                break
                        if not item_code_col:
                            result["errors"].append(f"Không tìm thấy cột Item Code trong {file}")
                            continue
                        # Lấy dữ liệu Item Code
                        for row in ws.iter_rows(min_row=11, min_col=item_code_col, max_col=item_code_col):
                            cell = row[0]
                            if cell.value:
                                text = str(cell.value).strip()
                                if text not in ["", "item code"]:
                                    all_values.add(text)
                    except Exception as e:
                        result["errors"].append(f"Lỗi đọc file {file}: {e}")
                # Ghi JSON
                output_data = {
                    "item_code": sorted(all_values)
                }
                with open(output_json, "w", encoding="utf-8") as f:
                    json.dump(output_data, f, ensure_ascii=False, indent=4)
                result["item_code_count"] = len(all_values)
                result["item_code_json"] = output_json
                result["item_code"] = sorted(all_values)

            # 10) Chạy trích barcode
            if downloaded_files:
                output_json = os.path.join(json_dir, "item_code.json")
                extract_item_code_to_json(downloaded_files, output_json)
            else:
                result["status"] = "no_xlsx_files"
    except Exception as exc:
        result["status"] = "error"
        result["errors"].append(str(exc))
    return result


if __name__ == "__main__":
    # Only run when executed directly. Importing this module will NOT trigger
    # network calls, so app.py can import the helpers safely.
    run_get_data_pog()